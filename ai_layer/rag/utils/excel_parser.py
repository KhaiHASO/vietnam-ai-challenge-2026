import io
import logging

logger = logging.getLogger(__name__)

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

class ExcelParser:
    @staticmethod
    def parse(file_content: bytes) -> str:
        """
        Parses text from a .xlsx file using openpyxl.
        Extracts row data into a formatted string.
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is not installed. Run `pip install openpyxl`.")
            
        try:
            workbook = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
            paragraphs = []
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                paragraphs.append(f"--- Sheet: {sheet_name} ---")
                
                for row in sheet.iter_rows(values_only=True):
                    # Filter out completely empty rows
                    if any(cell is not None for cell in row):
                        row_str = " | ".join(str(cell).strip() if cell is not None else "" for cell in row)
                        paragraphs.append(row_str)
                        
            return '\n'.join(paragraphs)
            
        except Exception as e:
            logger.error(f"Error parsing Excel: {e}")
            raise ValueError(f"Could not parse Excel file: {e}")
